"""Step 5: PDF and Excel report generation."""

import sqlite3
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from fpdf import FPDF
import openpyxl
from openpyxl.utils import get_column_letter

from db import queries
from db.queries import NOT_PARENT_SQL

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"


# ---------------------------------------------------------------------------
# Report modes
# ---------------------------------------------------------------------------
# Picked at report-run time. Controls which rows appear in the report body
# (the top banner always reflects the chosen scope so it's never ambiguous
# what was included).

MODE_FINALIZED = "finalized"
MODE_DRAFT = "draft"
MODE_ALL = "all"

MODE_CHOICES = [MODE_FINALIZED, MODE_DRAFT, MODE_ALL]
MODE_LABELS = {
    MODE_FINALIZED: "Finalized only",
    MODE_DRAFT: "Draft only",
    MODE_ALL: "All transactions",
}


def _filter_by_mode(rows, mode: str) -> list:
    """Filter a list of transaction rows by report mode."""
    if mode == MODE_FINALIZED:
        return [r for r in rows if r["status"] == "confirmed"]
    if mode == MODE_DRAFT:
        return [r for r in rows if r["status"] in ("pending", "needs_review")]
    return list(rows)  # MODE_ALL


# ---------------------------------------------------------------------------
# Data gathering helpers
# ---------------------------------------------------------------------------

def _fetch_month_transactions(conn: sqlite3.Connection, start: str, end: str,
                               mode: str = MODE_FINALIZED) -> list[dict]:
    """Fetch transactions for a date range, filtered by report mode.

    Split parents are excluded — their legs carry the dollars, so counting both
    would double the amount.
    """
    rows = queries.get_transactions(conn, start_date=start, end_date=end,
                                    exclude_parents=True)
    return [dict(r) for r in _filter_by_mode(rows, mode)]


def _fetch_ytd_transactions(conn: sqlite3.Connection, end: str,
                             mode: str = MODE_FINALIZED) -> list[dict]:
    """Fetch transactions from Jan 1 of the end-date year, filtered by mode."""
    year = end[:4]
    ytd_start = f"{year}-01-01"
    rows = queries.get_transactions(conn, start_date=ytd_start, end_date=end,
                                    exclude_parents=True)
    return [dict(r) for r in _filter_by_mode(rows, mode)]


def count_unfinalized(conn: sqlite3.Connection, start: str, end: str) -> tuple[int, Decimal]:
    """Count + sum of non-finalized (pending or needs_review) rows in range."""
    row = conn.execute(
        "SELECT COUNT(*) AS cnt, COALESCE(SUM(amount), 0) AS total "
        "FROM transactions WHERE date >= ? AND date <= ? "
        "AND status IN ('pending', 'needs_review') "
        f"AND {NOT_PARENT_SQL}",
        (start, end),
    ).fetchone()
    return row["cnt"], Decimal(str(row["total"]))


def count_missing_payee(conn: sqlite3.Connection, start: str, end: str,
                          mode: str = MODE_FINALIZED) -> tuple[int, Decimal]:
    """Count + sum of rows IN-SCOPE (mode-filtered) where payee IS NULL.

    Used for the missing-payee warning banner on each report.
    """
    if mode == MODE_FINALIZED:
        status_clause = "AND status = 'confirmed'"
    elif mode == MODE_DRAFT:
        status_clause = "AND status IN ('pending', 'needs_review')"
    else:  # MODE_ALL
        status_clause = ""
    sql = (
        f"SELECT COUNT(*) AS cnt, COALESCE(SUM(amount), 0) AS total "
        f"FROM transactions WHERE date >= ? AND date <= ? "
        f"AND payee IS NULL {status_clause} AND {NOT_PARENT_SQL}"
    )
    row = conn.execute(sql, (start, end)).fetchone()
    return row["cnt"], Decimal(str(row["total"]))


def _write_split_integrity(pdf: "AbacusPDF", conn) -> None:
    """Per-split balance check, rendered on page 1 below the checksums.

    This catches the one corruption the MATCH checksum cannot: a split whose
    legs no longer sum to its parent. The MATCH check excludes parents on both
    sides, so a missing/edited leg understates report and DB totals equally and
    still reads OK. This block verifies each split independently and names any
    that don't balance.
    """
    broken = queries.check_split_integrity(conn)
    pdf.section_title("Split Integrity")
    if not broken:
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(0, 5, "All splits balance (every split's legs sum to its parent).",
                 new_x="LMARGIN", new_y="NEXT")
        return

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_text_color(180, 0, 0)
    pdf.cell(0, 5,
             f"*** {len(broken)} UNBALANCED SPLIT(S) - inspect and fix in "
             f"Maintenance > Split Transaction ***",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    widths = [22, 26, 64, 28, 28, 28]
    pdf.table_header(widths, ["Parent ID", "Date", "Description",
                              "Parent", "Legs sum", "Difference"])
    for b in broken:
        pdf.table_row(
            widths,
            [str(b["id"]), b["date"], _safe(b["description_raw"])[:42],
             _fmt_amt(b["parent_amount"]), _fmt_amt(b["legs_total"]),
             _fmt_amt(b["difference"])],
            amount_cols=[3, 4, 5],
        )


def _sum_by(txns: list[dict], *keys) -> dict[tuple, Decimal]:
    """Group transactions by given keys and sum amounts."""
    totals: dict[tuple, Decimal] = {}
    for t in txns:
        group = tuple(t.get(k) or "" for k in keys)
        totals[group] = totals.get(group, Decimal(0)) + Decimal(t["amount"])
    return dict(sorted(totals.items()))


def _fmt_amt(val) -> str:
    """Format a Decimal/string amount as currency."""
    d = Decimal(str(val))
    if d >= 0:
        return f"${d:,.2f}"
    return f"-${abs(d):,.2f}"


def _safe(val) -> str:
    """Ensure a value is a clean string for PDF output."""
    if val is None:
        return ""
    s = str(val)
    s = s.replace("\u2014", "-")
    s = s.replace("\u2013", "-")
    s = s.replace("\u2018", "'")
    s = s.replace("\u2019", "'")
    s = s.replace("\u201c", '"')
    s = s.replace("\u201d", '"')
    s = s.replace("\u2026", "...")
    s = s.replace("&amp;", "&")
    return s.encode("latin-1", errors="replace").decode("latin-1")


# ---------------------------------------------------------------------------
# PDF helper class
# ---------------------------------------------------------------------------

class AbacusPDF(FPDF):
    """Landscape letter PDF with consistent headers/footers."""

    # GitHub-style green used for credits (positive amounts).
    GREEN_RGB = (63, 185, 80)

    def __init__(self, title: str = "Abacus Report"):
        super().__init__(orientation="L", unit="mm", format="Letter")
        self._title = _safe(title)
        self._section: str = ""  # set by section writers; included in per-page header
        self.set_auto_page_break(auto=True, margin=15)

    def set_section(self, section: str):
        """Set the section name shown in the per-page header.

        Call this right before pdf.add_page() so the new page's header picks
        up the new section. Carries over to continuation pages within the
        same section automatically.
        """
        self._section = _safe(section)

    def header(self):
        self.set_font("Helvetica", "B", 12)
        text = self._title
        if self._section:
            text = f"{text} — {self._section}"  # em dash separator
        self.cell(0, 8, _safe(text), new_x="LMARGIN", new_y="NEXT")
        self.ln(2)

    def footer(self):
        self.set_y(-10)
        self.set_font("Helvetica", "I", 7)
        self.cell(0, 5, f"Page {self.page_no()}/{{nb}}", align="C")

    def section_title(self, text: str):
        self.ln(4)
        self.set_font("Helvetica", "B", 10)
        self.cell(0, 6, _safe(text), new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(150, 150, 150)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.ln(2)

    def table_header(self, widths: list[float], headers: list[str]):
        self.set_font("Helvetica", "B", 7)
        self.set_fill_color(235, 235, 235)
        for w, h in zip(widths, headers):
            self.cell(w, 5, h, border=1, fill=True)
        self.ln()

    def table_row(self, widths: list[float], values: list[str],
                  bold: bool = False, fill: bool = False,
                  amount_cols: list[int] | None = None):
        """Print a row. If `amount_cols` lists column indices, those columns
        are interpreted as signed currency strings (e.g. "$12.34" or "-$5.00")
        and rendered in green when positive.
        """
        style = "B" if bold else ""
        self.set_font("Helvetica", style, 7)
        if fill:
            self.set_fill_color(245, 245, 245)
        amount_cols = set(amount_cols or [])
        for i, (w, v) in enumerate(zip(widths, values)):
            s = _safe(v)
            is_positive_amount = (
                i in amount_cols
                and s
                and not s.startswith("-")
                and ("$" in s)
            )
            if is_positive_amount:
                self.set_text_color(*self.GREEN_RGB)
                self.cell(w, 4.5, s, border=1, fill=fill, align="R")
                self.set_text_color(0, 0, 0)
            else:
                # Right-align money cols even when negative
                align = "R" if i in amount_cols else ""
                self.cell(w, 4.5, s, border=1, fill=fill, align=align)
        self.ln()

    def group_header(self, text: str, level: int = 0,
                     width: float | None = None):
        """Print a group header row. By default the grey bar spans the full
        printable width; pass `width` to limit it to the table's actual width.
        """
        indent = "    " * level
        self.set_font("Helvetica", "B", 7 if level > 0 else 8)
        if level == 0:
            self.set_fill_color(235, 235, 235)
        else:
            self.set_fill_color(242, 242, 242)
        bar_width = width if width is not None else (self.w - self.l_margin - self.r_margin)
        self.cell(bar_width, 5, f"{indent}{_safe(text)}", border=0, fill=True)
        self.ln()

    def subtotal_row(self, label: str, amount: str, indent: int = 0):
        """Print a subtotal row."""
        prefix = "    " * indent
        self.set_font("Helvetica", "B", 7)
        self.set_fill_color(248, 248, 248)
        w = self.w - self.l_margin - self.r_margin
        label_w = w - 35
        self.cell(label_w, 4.5, f"{prefix}{_safe(label)}", border=0, fill=True)
        self.cell(35, 4.5, _safe(amount), border=0, fill=True, align="R")
        self.ln()

    # -----------------------------------------------------------------------
    # Summary-table helpers (4-column: label | month-count | month-$ | YTD-count | YTD-$)
    # -----------------------------------------------------------------------

    def summary_header(self, label_w: float, count_w: float, amount_w: float,
                       label_title: str = "", has_ytd: bool = True):
        """Two-row table header for summary tables:
          Row 1: <label>  |  Month (spans 2)  |  YTD (spans 2)
          Row 2: Category |  Count |  Total   |  Count |  Total
        """
        self.set_font("Helvetica", "B", 7)
        self.set_fill_color(220, 220, 220)
        merged = count_w + amount_w
        # Row 1
        self.cell(label_w, 5, "", border=1, fill=True)
        self.cell(merged, 5, "Month", border=1, fill=True, align="C")
        if has_ytd:
            self.cell(merged, 5, "YTD", border=1, fill=True, align="C")
        self.ln()
        # Row 2
        self.set_fill_color(235, 235, 235)
        self.cell(label_w, 5, label_title, border=1, fill=True)
        self.cell(count_w, 5, "Count", border=1, fill=True, align="R")
        self.cell(amount_w, 5, "Total", border=1, fill=True, align="R")
        if has_ytd:
            self.cell(count_w, 5, "Count", border=1, fill=True, align="R")
            self.cell(amount_w, 5, "Total", border=1, fill=True, align="R")
        self.ln()

    def summary_row(self, label_w: float, count_w: float, amount_w: float,
                    label: str,
                    month_count: int | str, month_amt,
                    ytd_count: int | str | None = None, ytd_amt=None,
                    bold: bool = False, fill: bool = False):
        """Render a 4-col data row (or 2-col when ytd_count is None).

        Amount columns auto-render in green when the underlying value is > 0.
        """
        style = "B" if bold else ""
        self.set_font("Helvetica", style, 7)
        if fill:
            self.set_fill_color(245, 245, 245)

        # Label
        self.cell(label_w, 4.5, _safe(label), border=1, fill=fill)

        # Month count + amount
        self.cell(count_w, 4.5, _safe(str(month_count) if month_count != "" else ""),
                  border=1, fill=fill, align="R")
        self._amount_cell(amount_w, month_amt, fill=fill)

        # YTD count + amount (optional)
        if ytd_count is not None:
            self.cell(count_w, 4.5, _safe(str(ytd_count) if ytd_count != "" else ""),
                      border=1, fill=fill, align="R")
            self._amount_cell(amount_w, ytd_amt, fill=fill)

        self.ln()

    def summary_subtotal_header(self, label_w: float, count_w: float, amount_w: float,
                                  label: str,
                                  month_count: int | str, month_amt,
                                  ytd_count: int | str | None = None, ytd_amt=None,
                                  level: int = 0):
        """A banner row that doubles as a subtotal: grey-fill row with the
        group label on the left and bold count/amount cells on the right that
        match the data-table column widths. Used as a category or subcategory
        header that simultaneously shows the subtotal — avoids a separate
        "SUBTOTAL" line below the group.

        level=0 gets a darker grey (matches existing group_header level 0);
        level=1 gets the lighter grey for subcategory bands.
        """
        indent = "    " * level
        fill_color = (235, 235, 235) if level == 0 else (242, 242, 242)
        self.set_fill_color(*fill_color)
        self.set_font("Helvetica", "B", 8 if level == 0 else 7)

        self.cell(label_w, 5, f"{indent}{_safe(label)}", border=1, fill=True)
        self.cell(count_w, 5,
                  _safe(str(month_count) if month_count != "" else ""),
                  border=1, fill=True, align="R")
        self._amount_cell(amount_w, month_amt, fill=True, height=5)
        if ytd_count is not None:
            self.cell(count_w, 5,
                      _safe(str(ytd_count) if ytd_count != "" else ""),
                      border=1, fill=True, align="R")
            self._amount_cell(amount_w, ytd_amt, fill=True, height=5)
        self.ln()

    def _amount_cell(self, width: float, amount, fill: bool = False,
                     height: float = 4.5):
        """Render one currency cell. Green text when amount > 0; black otherwise.
        Accepts Decimal, float, int, or pre-formatted string.
        """
        # Determine numeric value to decide color
        is_positive = False
        try:
            if isinstance(amount, str):
                if amount and not amount.startswith("-") and "$" in amount:
                    is_positive = True
            else:
                is_positive = Decimal(str(amount)) > 0
        except Exception:
            is_positive = False

        # Format if numeric, otherwise treat as pre-formatted
        if isinstance(amount, str):
            text = amount
        elif amount is None or amount == "":
            text = ""
        else:
            text = _fmt_amt(amount)

        if is_positive:
            self.set_text_color(*self.GREEN_RGB)
        self.cell(width, height, _safe(text), border=1, fill=fill, align="R")
        self.set_text_color(0, 0, 0)


# ---------------------------------------------------------------------------
# Section builders
# ---------------------------------------------------------------------------

def _category_sort_key(cat: str) -> tuple[int, str]:
    """Sort key that puts Income at the top, Transfer at the bottom, and
    every other category in between (alphabetical). Used to give the
    Category Summary and Payee Summary the same top-down flow as the
    Cash Flow Summary on page 2: income sources first, then spending."""
    if cat == "Income":
        return (0, cat)
    if cat == "Transfer":
        return (2, cat)
    return (1, cat)

def _write_cash_flow_summary(pdf: AbacusPDF, month_txns, ytd_txns):
    """Cash flow summary: breaks Income into its subcategories, then computes
    'Burn Rate' as the sum of everything that's not Income and not Transfer.
    Credits in non-Income categories (refunds, returns, tax refunds) naturally
    offset debits in the same category because the sum nets them. Final line
    shows Net Cash Flow = Income + Burn (where Burn is negative).

    Inputs may contain Transfer transactions; this function filters them out.
    """
    pdf.section_title("Cash Flow Summary (Month & YTD)")

    INCOME = "Income"
    EXCLUDED_FROM_BURN = {"Income", "Transfer"}

    def _aggregate(txns):
        """Return (income_by_subcat, burn_count, burn_total)."""
        income: dict[str, list] = {}  # subcat -> [count, total]
        burn_c, burn_t = 0, Decimal(0)
        for t in txns:
            cat = t.get("category") or ""
            sub = t.get("subcategory") or ""
            amt = Decimal(str(t["amount"]))
            if cat == INCOME:
                slot = income.setdefault(sub, [0, Decimal(0)])
                slot[0] += 1
                slot[1] += amt
            elif cat in EXCLUDED_FROM_BURN:
                continue  # exclude transfers
            else:
                burn_c += 1
                burn_t += amt
        return income, burn_c, burn_t

    m_income, m_burn_c, m_burn_t = _aggregate(month_txns)
    y_income, y_burn_c, y_burn_t = _aggregate(ytd_txns)

    LABEL_W, COUNT_W, AMT_W = 100, 20, 35
    table_w = LABEL_W + (COUNT_W + AMT_W) * 2
    pdf.summary_header(LABEL_W, COUNT_W, AMT_W, "Item")

    # Income breakdown
    pdf.group_header("Income (sources of cash)", level=0, width=table_w)
    income_subs = sorted(set(m_income.keys()) | set(y_income.keys()))
    for sub in income_subs:
        m_c, m_t = m_income.get(sub, [0, Decimal(0)])
        y_c, y_t = y_income.get(sub, [0, Decimal(0)])
        label = f"    {sub or '(no subcategory)'}"
        pdf.summary_row(LABEL_W, COUNT_W, AMT_W,
                        label,
                        m_c or "", m_t, y_c or "", y_t)
    # Income subtotal
    m_inc_c = sum(slot[0] for slot in m_income.values())
    m_inc_t = sum((slot[1] for slot in m_income.values()), Decimal(0))
    y_inc_c = sum(slot[0] for slot in y_income.values())
    y_inc_t = sum((slot[1] for slot in y_income.values()), Decimal(0))
    pdf.summary_row(LABEL_W, COUNT_W, AMT_W,
                    "  TOTAL INCOME",
                    m_inc_c or "", m_inc_t, y_inc_c or "", y_inc_t,
                    bold=True, fill=True)

    pdf.ln(2)

    # Burn rate
    pdf.group_header("Spending (Burn Rate)", level=0, width=table_w)
    pdf.summary_row(LABEL_W, COUNT_W, AMT_W,
                    "    Net spending (all categories except Income / Transfer;",
                    m_burn_c, m_burn_t, y_burn_c, y_burn_t,
                    bold=True, fill=True)
    pdf.set_font("Helvetica", "I", 6)
    pdf.cell(0, 4, "    credits like refunds/returns/tax refunds offset debits within the same category)")
    pdf.ln()
    pdf.set_font("Helvetica", "", 7)

    pdf.ln(2)

    # Net cash flow = Income + Burn (burn is negative, so this is income - spending)
    m_net = m_inc_t + m_burn_t
    y_net = y_inc_t + y_burn_t
    pdf.summary_row(LABEL_W, COUNT_W, AMT_W,
                    "NET CASH FLOW (Income + Burn)",
                    "", m_net, "", y_net,
                    bold=True, fill=True)

    pdf.ln(6)


def _write_category_summary(pdf: AbacusPDF, month_txns, ytd_txns):
    """Section 1: Category Summary (Month & YTD) - grouped by category.

    Four data columns: month count, month total, YTD count, YTD total.
    Credits (positive totals) render in green.
    Prefixed by a Cash Flow Summary (income vs burn rate vs net).
    """
    # Cash flow summary at the top of this section
    _write_cash_flow_summary(pdf, month_txns, ytd_txns)

    pdf.section_title("Category Summary (Month & YTD)")

    def _aggregate(txns):
        """Return (count_by_key, total_by_key) where key is (cat, subcat)."""
        counts: dict[tuple, int] = {}
        totals: dict[tuple, Decimal] = {}
        for t in txns:
            key = (t.get("category") or "", t.get("subcategory") or "")
            counts[key] = counts.get(key, 0) + 1
            totals[key] = totals.get(key, Decimal(0)) + Decimal(t["amount"])
        return counts, totals

    m_counts, m_totals = _aggregate(month_txns)
    y_counts, y_totals = _aggregate(ytd_txns)
    all_keys = sorted(set(m_totals.keys()) | set(y_totals.keys()))

    # Unique categories, ordered: Income first, others alphabetical, Transfer last
    categories: list[str] = sorted(
        {k[0] for k in all_keys},
        key=_category_sort_key,
    )

    # Column widths: label, month count, month $, ytd count, ytd $
    LABEL_W, COUNT_W, AMT_W = 100, 20, 35
    table_w = LABEL_W + (COUNT_W + AMT_W) * 2
    pdf.summary_header(LABEL_W, COUNT_W, AMT_W, "Category")

    for cat in categories:
        # Category banner row doubles as the subtotal — shows the category
        # name on the left and its month/YTD totals on the right in bold.
        cat_m_c = sum(v for k, v in m_counts.items() if k[0] == cat)
        cat_m_t = sum((v for k, v in m_totals.items() if k[0] == cat), Decimal(0))
        cat_y_c = sum(v for k, v in y_counts.items() if k[0] == cat)
        cat_y_t = sum((v for k, v in y_totals.items() if k[0] == cat), Decimal(0))
        pdf.summary_subtotal_header(LABEL_W, COUNT_W, AMT_W,
                                     cat,
                                     cat_m_c or "", cat_m_t,
                                     cat_y_c or "", cat_y_t,
                                     level=0)

        # Subcategory data rows below the banner
        cat_keys = [k for k in all_keys if k[0] == cat]
        for key in cat_keys:
            _, subcat = key
            if subcat:
                m_c = m_counts.get(key, 0) or ""
                m_t = m_totals.get(key, Decimal(0))
                y_c = y_counts.get(key, 0) or ""
                y_t = y_totals.get(key, Decimal(0))
                pdf.summary_row(LABEL_W, COUNT_W, AMT_W,
                                f"    {subcat}",
                                m_c, m_t, y_c, y_t)

    # Grand total
    grand_m_c = sum(m_counts.values())
    grand_m_t = sum(m_totals.values(), Decimal(0))
    grand_y_c = sum(y_counts.values())
    grand_y_t = sum(y_totals.values(), Decimal(0))
    pdf.ln(2)
    pdf.summary_row(LABEL_W, COUNT_W, AMT_W,
                    "GRAND TOTAL",
                    grand_m_c, grand_m_t, grand_y_c, grand_y_t,
                    bold=True, fill=True)


def _write_payee_summary(pdf: AbacusPDF, month_txns, ytd_txns):
    """Section 2: Payee Summary (Month & YTD) - grouped by category/subcategory.

    Four data columns: month count, month total, YTD count, YTD total.
    Credits (positive totals) render in green.
    """
    pdf.section_title("Payee Summary (Month & YTD)")

    def _aggregate_by_payee(txns):
        """{(cat, subcat, payee): (count, total)}"""
        out: dict[tuple, tuple[int, Decimal]] = {}
        for t in txns:
            key = (t.get("category") or "", t.get("subcategory") or "",
                   t.get("payee") or "")
            c, tot = out.get(key, (0, Decimal(0)))
            out[key] = (c + 1, tot + Decimal(t["amount"]))
        return out

    m_by_payee = _aggregate_by_payee(month_txns)
    y_by_payee = _aggregate_by_payee(ytd_txns)
    # Sort: Income categories first (alphabetical within), then other expense
    # categories alphabetical, then Transfer last. Within each category,
    # alphabetical by subcategory and payee.
    all_keys = sorted(
        set(m_by_payee.keys()) | set(y_by_payee.keys()),
        key=lambda k: (_category_sort_key(k[0]), k[1], k[2]),
    )

    LABEL_W, COUNT_W, AMT_W = 100, 20, 35
    pdf.summary_header(LABEL_W, COUNT_W, AMT_W, "Payee")

    # Pre-compute subtotals at the category and (category, subcategory) levels
    # so the banner rows can show them inline.
    def _subtotals_at(level: int):
        """level=1 -> by (cat,), level=2 -> by (cat, subcat)."""
        m_out: dict[tuple, tuple[int, Decimal]] = {}
        y_out: dict[tuple, tuple[int, Decimal]] = {}
        for src, dest in [(m_by_payee, m_out), (y_by_payee, y_out)]:
            for k, (c, t) in src.items():
                bucket = k[:level]
                pc, pt = dest.get(bucket, (0, Decimal(0)))
                dest[bucket] = (pc + c, pt + t)
        return m_out, y_out

    m_cat, y_cat = _subtotals_at(1)
    m_subcat, y_subcat = _subtotals_at(2)

    current_cat = None
    current_subcat = None

    for key in all_keys:
        cat, subcat, payee = key
        m_c, m_t = m_by_payee.get(key, (0, Decimal(0)))
        y_c, y_t = y_by_payee.get(key, (0, Decimal(0)))

        if cat != current_cat:
            cm_c, cm_t = m_cat.get((cat,), (0, Decimal(0)))
            cy_c, cy_t = y_cat.get((cat,), (0, Decimal(0)))
            pdf.summary_subtotal_header(LABEL_W, COUNT_W, AMT_W,
                                          cat,
                                          cm_c or "", cm_t,
                                          cy_c or "", cy_t,
                                          level=0)
            current_cat = cat
            current_subcat = None

        if subcat and subcat != current_subcat:
            sm_c, sm_t = m_subcat.get((cat, subcat), (0, Decimal(0)))
            sy_c, sy_t = y_subcat.get((cat, subcat), (0, Decimal(0)))
            pdf.summary_subtotal_header(LABEL_W, COUNT_W, AMT_W,
                                          subcat,
                                          sm_c or "", sm_t,
                                          sy_c or "", sy_t,
                                          level=1)
            current_subcat = subcat

        indent = "        " if subcat else "    "
        pdf.summary_row(LABEL_W, COUNT_W, AMT_W,
                        f"{indent}{payee}",
                        m_c or "", m_t,
                        y_c or "", y_t)

    # Grand total
    grand_m_c = sum(c for c, _ in m_by_payee.values())
    grand_m_t = sum((t for _, t in m_by_payee.values()), Decimal(0))
    grand_y_c = sum(c for c, _ in y_by_payee.values())
    grand_y_t = sum((t for _, t in y_by_payee.values()), Decimal(0))
    pdf.ln(2)
    pdf.summary_row(LABEL_W, COUNT_W, AMT_W,
                    "GRAND TOTAL",
                    grand_m_c, grand_m_t, grand_y_c, grand_y_t,
                    bold=True, fill=True)


def _write_transaction_detail(pdf: AbacusPDF, month_txns):
    """Section 3: Transaction Detail (Month) - grouped by category/subcategory.

    Credit amounts render in green; grey group headers span only the table width.
    """
    pdf.section_title("Transaction Detail (Month)")

    sorted_txns = sorted(month_txns, key=lambda x: (
        x.get("category") or "", x.get("subcategory") or "", x["date"]
    ))

    widths = [20, 50, 20, 25, 20, 40]
    table_w = sum(widths)
    pdf.table_header(widths, ["Date", "Payee", "Via", "Amount", "Payor", "Note"])

    current_cat = None
    current_subcat = None

    for t in sorted_txns:
        cat = t.get("category") or ""
        subcat = t.get("subcategory") or ""

        if cat != current_cat:
            pdf.group_header(cat, level=0, width=table_w)
            current_cat = cat
            current_subcat = None

        if subcat and subcat != current_subcat:
            pdf.group_header(subcat, level=1, width=table_w)
            current_subcat = subcat

        pdf.table_row(widths, [
            t["date"], t.get("payee") or "", t.get("via") or "",
            _fmt_amt(t["amount"]), t.get("payor") or "", t.get("note") or "",
        ], amount_cols=[3])


def _write_tax_items(pdf: AbacusPDF, month_txns, ytd_txns):
    """Section 4: Tax Items Report (Month & YTD) - grouped by tax flag.

    Starts with a 4-column summary table (flag | month count | month $ | YTD
    count | YTD $), then lists per-flag detail with month transactions.
    Credits (positive amounts) render in green throughout.
    """
    pdf.section_title("Tax Items (Month & YTD)")

    def _by_flag(txns):
        by_flag: dict[str, list] = {}
        for t in txns:
            if not t.get("tax_flags"):
                continue
            for flag in t["tax_flags"].split(","):
                flag = flag.strip()
                if flag:
                    by_flag.setdefault(flag, []).append(t)
        return by_flag

    month_by_flag = _by_flag(month_txns)
    ytd_by_flag = _by_flag(ytd_txns)
    all_flags = sorted(set(month_by_flag.keys()) | set(ytd_by_flag.keys()))

    if not all_flags:
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(0, 6, "No tax-flagged transactions in this period.")
        return

    # --- 4-column summary table at top ---
    LABEL_W, COUNT_W, AMT_W = 100, 20, 35
    pdf.summary_header(LABEL_W, COUNT_W, AMT_W, "Tax Flag")
    for flag in all_flags:
        m_t = month_by_flag.get(flag, [])
        y_t = ytd_by_flag.get(flag, [])
        pdf.summary_row(LABEL_W, COUNT_W, AMT_W,
                        flag,
                        len(m_t), sum((Decimal(t["amount"]) for t in m_t), Decimal(0)),
                        len(y_t), sum((Decimal(t["amount"]) for t in y_t), Decimal(0)))

    pdf.ln(4)

    # --- Per-flag detail listings ---
    detail_widths = [20, 50, 40, 25, 40]
    detail_table_w = sum(detail_widths)

    for flag in all_flags:
        pdf.group_header(f"{flag} — Detail", level=0, width=detail_table_w)
        pdf.table_header(detail_widths, ["Date", "Payee", "Category", "Amount", "Note"])

        m_txns = month_by_flag.get(flag, [])
        y_txns = ytd_by_flag.get(flag, [])

        for t in sorted(m_txns, key=lambda x: x["date"]):
            cat_label = t.get("category") or ""
            if t.get("subcategory"):
                cat_label += f" / {t['subcategory']}"
            pdf.table_row(detail_widths, [
                t["date"], t.get("payee") or "", cat_label,
                _fmt_amt(t["amount"]), t.get("note") or "",
            ], amount_cols=[3])

        m_total = sum((Decimal(t["amount"]) for t in m_txns), Decimal(0))
        y_total = sum((Decimal(t["amount"]) for t in y_txns), Decimal(0))
        pdf.subtotal_row(f"Month subtotal - {flag}", _fmt_amt(m_total), indent=1)
        pdf.subtotal_row(f"YTD subtotal - {flag}", _fmt_amt(y_total), indent=1)
        pdf.ln(2)


def _write_checksums(pdf: AbacusPDF, conn, start, end,
                     month_txns, ytd_txns,
                     month_no_xfer, ytd_no_xfer, month_xfer,
                     mode: str = MODE_FINALIZED):
    """Checksums section - cross-reference counts and totals for integrity.

    The 'in-scope' total compares against the report's selected mode rather
    than always against 'confirmed'. The MATCH check passes when the report
    body's count and dollar total equal the DB total for the mode-filtered
    rows in the period.
    """
    mode_label = MODE_LABELS.get(mode, mode)
    pdf.section_title(f"Checksums - Data Integrity Verification ({mode_label})")

    year = end[:4]
    ytd_start = f"{year}-01-01"

    # Database counts (all statuses). Split parents are excluded from every
    # checksum query so that all numbers here reflect counted rows and the
    # MATCH check lines up with the report body (which also excludes parents).
    db_month_all = conn.execute(
        "SELECT COUNT(*) as cnt, COALESCE(SUM(amount), 0) as total "
        f"FROM transactions WHERE date >= ? AND date <= ? AND {NOT_PARENT_SQL}",
        (start, end),
    ).fetchone()

    db_ytd_all = conn.execute(
        "SELECT COUNT(*) as cnt, COALESCE(SUM(amount), 0) as total "
        f"FROM transactions WHERE date >= ? AND date <= ? AND {NOT_PARENT_SQL}",
        (ytd_start, end),
    ).fetchone()

    # Database counts (in scope for the chosen mode)
    if mode == MODE_FINALIZED:
        scope_clause = "status = 'confirmed'"
        out_of_scope_clause = "status IN ('pending', 'needs_review')"
    elif mode == MODE_DRAFT:
        scope_clause = "status IN ('pending', 'needs_review')"
        out_of_scope_clause = "status = 'confirmed'"
    else:  # MODE_ALL
        scope_clause = "1=1"
        out_of_scope_clause = "0=1"

    db_month_in_scope = conn.execute(
        f"SELECT COUNT(*) as cnt, COALESCE(SUM(amount), 0) as total "
        f"FROM transactions WHERE date >= ? AND date <= ? AND {scope_clause} "
        f"AND {NOT_PARENT_SQL}",
        (start, end),
    ).fetchone()

    db_ytd_in_scope = conn.execute(
        f"SELECT COUNT(*) as cnt, COALESCE(SUM(amount), 0) as total "
        f"FROM transactions WHERE date >= ? AND date <= ? AND {scope_clause} "
        f"AND {NOT_PARENT_SQL}",
        (ytd_start, end),
    ).fetchone()

    # Counts outside the chosen scope (just for reference)
    db_month_out_of_scope = conn.execute(
        f"SELECT COUNT(*) as cnt FROM transactions "
        f"WHERE date >= ? AND date <= ? AND {out_of_scope_clause} "
        f"AND {NOT_PARENT_SQL}",
        (start, end),
    ).fetchone()["cnt"]

    db_ytd_out_of_scope = conn.execute(
        f"SELECT COUNT(*) as cnt FROM transactions "
        f"WHERE date >= ? AND date <= ? AND {out_of_scope_clause} "
        f"AND {NOT_PARENT_SQL}",
        (ytd_start, end),
    ).fetchone()["cnt"]

    # Report totals (what's actually in the report sections)
    rpt_month_count = len(month_txns)
    rpt_month_total = sum(Decimal(t["amount"]) for t in month_txns)
    rpt_month_no_xfer_count = len(month_no_xfer)
    rpt_month_no_xfer_total = sum(Decimal(t["amount"]) for t in month_no_xfer)
    rpt_xfer_count = len(month_xfer)
    rpt_xfer_total = sum(Decimal(t["amount"]) for t in month_xfer)

    rpt_ytd_count = len(ytd_txns)
    rpt_ytd_total = sum(Decimal(t["amount"]) for t in ytd_txns)
    rpt_ytd_no_xfer_count = len(ytd_no_xfer)
    rpt_ytd_no_xfer_total = sum(Decimal(t["amount"]) for t in ytd_no_xfer)

    # By source
    month_by_source = {}
    for t in month_txns:
        src = t.get("source") or "Unknown"
        if src not in month_by_source:
            month_by_source[src] = {"count": 0, "total": Decimal(0)}
        month_by_source[src]["count"] += 1
        month_by_source[src]["total"] += Decimal(t["amount"])

    # Render
    widths = [100, 25, 35]
    pdf.table_header(widths, ["Metric", "Count", "Amount"])

    in_scope_label = f"Database in scope ({mode_label})"
    out_of_scope_label = "Database out of scope" if mode != MODE_ALL else "(no out-of-scope rows)"

    pdf.group_header(f"Month: {start} to {end}", level=0)
    pdf.table_row(widths, ["Database total (all statuses)",
                           str(db_month_all["cnt"]),
                           _fmt_amt(db_month_all["total"])])
    pdf.table_row(widths, [in_scope_label,
                           str(db_month_in_scope["cnt"]),
                           _fmt_amt(db_month_in_scope["total"])])
    pdf.table_row(widths, [out_of_scope_label,
                           str(db_month_out_of_scope), ""])
    pdf.table_row(widths, ["Report total (in scope, incl transfers)",
                           str(rpt_month_count),
                           _fmt_amt(rpt_month_total)])
    pdf.table_row(widths, ["Report excl transfers (used in summaries)",
                           str(rpt_month_no_xfer_count),
                           _fmt_amt(rpt_month_no_xfer_total)])
    pdf.table_row(widths, ["Transfers excluded",
                           str(rpt_xfer_count),
                           _fmt_amt(rpt_xfer_total)])

    # Match check — quantize to cents before comparing. SQLite stores `amount`
    # as REAL (float), so the DB-side SUM and the Python-side Decimal sum can
    # diverge by ~1e-13 from float noise even though they round to the same
    # dollar value. The user-visible totals are already in cents; equality
    # at that precision is the correct integrity signal.
    def _cents(v):
        return Decimal(str(v)).quantize(Decimal("0.01"))
    month_match = (rpt_month_count == db_month_in_scope["cnt"] and
                   _cents(rpt_month_total) == _cents(db_month_in_scope["total"]))
    pdf.table_row(widths, [
        "MONTH MATCH" if month_match else "*** MONTH MISMATCH ***",
        "OK" if month_match else "FAIL", ""
    ], bold=True, fill=True)

    pdf.ln(3)
    pdf.group_header(f"YTD: {ytd_start} to {end}", level=0)
    pdf.table_row(widths, ["Database total (all statuses)",
                           str(db_ytd_all["cnt"]),
                           _fmt_amt(db_ytd_all["total"])])
    pdf.table_row(widths, [in_scope_label,
                           str(db_ytd_in_scope["cnt"]),
                           _fmt_amt(db_ytd_in_scope["total"])])
    pdf.table_row(widths, [out_of_scope_label,
                           str(db_ytd_out_of_scope), ""])
    pdf.table_row(widths, ["Report total (in scope, incl transfers)",
                           str(rpt_ytd_count),
                           _fmt_amt(rpt_ytd_total)])
    pdf.table_row(widths, ["Report excl transfers",
                           str(rpt_ytd_no_xfer_count),
                           _fmt_amt(rpt_ytd_no_xfer_total)])

    ytd_match = (rpt_ytd_count == db_ytd_in_scope["cnt"] and
                 _cents(rpt_ytd_total) == _cents(db_ytd_in_scope["total"]))
    pdf.table_row(widths, [
        "YTD MATCH" if ytd_match else "*** YTD MISMATCH ***",
        "OK" if ytd_match else "FAIL", ""
    ], bold=True, fill=True)

    # By source breakdown
    pdf.ln(3)
    pdf.group_header("Month by Source", level=0)
    src_widths = [80, 25, 35]
    pdf.table_header(src_widths, ["Source", "Count", "Total"])
    for src in sorted(month_by_source.keys()):
        d = month_by_source[src]
        pdf.table_row(src_widths, [src, str(d["count"]), _fmt_amt(d["total"])])


# ---------------------------------------------------------------------------
# Public report generators
# ---------------------------------------------------------------------------

def _write_top_banners(pdf: AbacusPDF, conn: sqlite3.Connection,
                        start: str, end: str, mode: str):
    """Top-of-report scope banner + missing-payee warning."""
    nf_count, nf_total = count_unfinalized(conn, start, end)
    mp_count, mp_total = count_missing_payee(conn, start, end, mode)

    # Scope label
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, f"Scope: {MODE_LABELS[mode]}",
             new_x="LMARGIN", new_y="NEXT")

    # Mode-specific inclusion/exclusion line
    pdf.set_font("Helvetica", "", 9)
    if mode == MODE_FINALIZED:
        banner = (f"This report EXCLUDES {nf_count} transaction(s) totaling "
                  f"{_fmt_amt(nf_total)} that are not finalized.")
    elif mode == MODE_DRAFT:
        banner = (f"This report INCLUDES ONLY draft transactions: {nf_count} "
                  f"totaling {_fmt_amt(nf_total)} (not yet finalized).")
    else:  # MODE_ALL
        banner = (f"This report INCLUDES {nf_count} transaction(s) totaling "
                  f"{_fmt_amt(nf_total)} that are not finalized.")
    pdf.cell(0, 5, _safe(banner), new_x="LMARGIN", new_y="NEXT")

    # Missing-payee warning (only if any in-scope rows have NULL payee)
    if mp_count:
        pdf.set_font("Helvetica", "BI", 9)
        warning = (f"WARNING: This report includes {mp_count} transaction(s) "
                   f"totaling {_fmt_amt(mp_total)} where the payee is missing.")
        pdf.cell(0, 5, _safe(warning), new_x="LMARGIN", new_y="NEXT")

    pdf.ln(3)


def generate_monthly_pdf(conn: sqlite3.Connection, start: str, end: str,
                         title: str | None = None,
                         sections: list[str] | None = None,
                         mode: str = MODE_FINALIZED) -> Path:
    """Generate the monthly PDF report.

    sections controls which sections to include. Options:
    'category_summary', 'payee_summary', 'transaction_detail', 'tax_items'
    If None, all sections are included.

    mode: one of MODE_FINALIZED (default), MODE_DRAFT, or MODE_ALL.
    Controls which transactions appear in the body. The top banner always
    reflects the chosen mode.
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    if sections is None:
        sections = ["category_summary", "payee_summary", "transaction_detail", "tax_items"]

    month_txns = _fetch_month_transactions(conn, start, end, mode=mode)
    ytd_txns = _fetch_ytd_transactions(conn, end, mode=mode)

    # Separate transfers
    month_no_xfer = [t for t in month_txns if t.get("category") != "Transfer"]
    ytd_no_xfer = [t for t in ytd_txns if t.get("category") != "Transfer"]
    month_xfer = [t for t in month_txns if t.get("category") == "Transfer"]

    report_title = title or f"Monthly Report - {start} to {end}"
    pdf = AbacusPDF(title=report_title)
    pdf.alias_nb_pages()

    # Page 1: scope + missing-payee banners, then the checksum table.
    pdf.set_section("Scope & Checksums")
    pdf.add_page()
    _write_top_banners(pdf, conn, start, end, mode)
    _write_checksums(pdf, conn, start, end, month_txns, ytd_txns,
                     month_no_xfer, ytd_no_xfer, month_xfer, mode=mode)
    _write_split_integrity(pdf, conn)

    if "category_summary" in sections:
        pdf.set_section("Category Summary")
        pdf.add_page()
        _write_category_summary(pdf, month_no_xfer, ytd_no_xfer)
        if month_xfer:
            xfer_total = sum(Decimal(t["amount"]) for t in month_xfer)
            pdf.set_font("Helvetica", "I", 7)
            pdf.ln(2)
            pdf.cell(0, 4, f"Note: {len(month_xfer)} transfer(s) totaling {_fmt_amt(xfer_total)} excluded from summaries.")

    if "payee_summary" in sections:
        pdf.set_section("Payee Summary")
        pdf.add_page()
        _write_payee_summary(pdf, month_no_xfer, ytd_no_xfer)

    if "transaction_detail" in sections:
        pdf.set_section("Transaction Detail")
        pdf.add_page()
        _write_transaction_detail(pdf, month_txns)

    if "tax_items" in sections:
        pdf.set_section("Tax Items")
        pdf.add_page()
        _write_tax_items(pdf, month_no_xfer, ytd_no_xfer)

    filename = f"Monthly_Report_{start}_to_{end}.pdf"
    out_path = OUTPUT_DIR / filename
    pdf.output(str(out_path))
    return out_path


def generate_pending_items_pdf(conn: sqlite3.Connection) -> Path | None:
    """Generate a list of all pending transactions, sorted by payee."""
    OUTPUT_DIR.mkdir(exist_ok=True)

    pending = queries.get_pending_transactions(conn)
    if not pending:
        return None

    # Sort by payee, then date
    sorted_pending = sorted(pending, key=lambda t: (t["payee"] or "", t["date"]))

    pdf = AbacusPDF(title=f"Pending Items - {len(sorted_pending)} transaction(s)")
    pdf.alias_nb_pages()
    pdf.add_page()

    widths = [40, 20, 28, 30, 100, 40]
    pdf.table_header(widths, ["Payee", "Date", "Amount", "Source", "Description", "Note"])

    for t in sorted_pending:
        pdf.table_row(widths, [
            t["payee"] or "", t["date"], _fmt_amt(t["amount"]),
            t["source"], t["description_raw"], t["note"] or "",
        ])

    # Timestamp the filename so re-runs don't collide with an open copy of the
    # previous PDF (Windows locks files that any viewer has open).
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    out_path = OUTPUT_DIR / f"Pending_Items_{timestamp}.pdf"
    pdf.output(str(out_path))
    return out_path


def generate_excel_export(conn: sqlite3.Connection, start: str, end: str,
                          mode: str = MODE_FINALIZED) -> Path:
    """Generate an Excel export of transactions in a date range, filtered by mode."""
    OUTPUT_DIR.mkdir(exist_ok=True)

    rows = queries.get_transactions(conn, start_date=start, end_date=end,
                                    exclude_parents=True)
    rows = _filter_by_mode(rows, mode)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Transactions"

    headers = [
        "Date", "Amount", "Check Number", "Description (Raw)", "Category (Raw)",
        "Payee", "Via", "Payor", "Category", "Subcategory", "Tax Flags",
        "Note", "Source", "Status", "Overridden",
    ]
    field_keys = [
        "date", "amount", "check_number", "description_raw", "category_raw",
        "payee", "via", "payor", "category", "subcategory", "tax_flags",
        "note", "source", "status", "overridden",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    for row_idx, txn in enumerate(rows, 2):
        for col_idx, key in enumerate(field_keys, 1):
            val = txn[key]
            if key == "amount":
                val = float(val) if val else 0.0
            elif key == "overridden":
                val = bool(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(h) + 4, 12)

    mode_suffix = {MODE_FINALIZED: "finalized", MODE_DRAFT: "draft", MODE_ALL: "all"}.get(mode, mode)
    filename = f"Transactions_{start}_to_{end}_{mode_suffix}.xlsx"
    out_path = OUTPUT_DIR / filename
    wb.save(str(out_path))
    return out_path
