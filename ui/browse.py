"""Browse / Search — view and filter all transactions."""

import io
import streamlit as st
import pandas as pd
from datetime import date, timedelta

import openpyxl
from openpyxl.utils import get_column_letter

from db import queries


def browse_page(conn):
    st.title("Browse / Search")

    # --- Search bars ---
    col_s1, col_s2 = st.columns(2)
    search_payee = col_s1.text_input("Search in Payee", placeholder="e.g. Safeway")
    search_all = col_s2.text_input("Search Anywhere", placeholder="e.g. Medical, Venmo, Chase...")

    # --- Date range filters ---
    col1, col2, col3 = st.columns([1, 1, 1])

    with col1:
        preset = st.selectbox("Date preset", [
            "All time", "This month", "Last month", "This quarter",
            "YTD", "Last year", "Custom",
        ])

    today = date.today()
    if preset == "All time":
        start_val = None
        end_val = None
    elif preset == "This month":
        start_val = today.replace(day=1)
        end_val = today
    elif preset == "Last month":
        first_this_month = today.replace(day=1)
        last_month_end = first_this_month - timedelta(days=1)
        start_val = last_month_end.replace(day=1)
        end_val = last_month_end
    elif preset == "This quarter":
        q_month = ((today.month - 1) // 3) * 3 + 1
        start_val = today.replace(month=q_month, day=1)
        end_val = today
    elif preset == "YTD":
        start_val = today.replace(month=1, day=1)
        end_val = today
    elif preset == "Last year":
        start_val = today.replace(year=today.year - 1, month=1, day=1)
        end_val = today.replace(year=today.year - 1, month=12, day=31)
    else:  # Custom
        start_val = today.replace(month=1, day=1)
        end_val = today

    with col2:
        start_date = st.date_input("From", value=start_val, key="browse_start",
                                   disabled=(preset != "Custom" and preset != "All time"))
    with col3:
        end_date = st.date_input("To", value=end_val, key="browse_end",
                                 disabled=(preset != "Custom" and preset != "All time"))

    # --- Source and status filters ---
    col4, col5 = st.columns(2)
    sources = queries.get_distinct_sources(conn)
    source_filter = col4.selectbox("Source", ["All"] + sources, key="browse_source")
    status_filter = col5.selectbox("Status", ["All", "pending", "confirmed", "needs_review"],
                                   key="browse_status")

    # --- Query ---
    rows = queries.get_transactions(
        conn,
        start_date=start_date.isoformat() if start_date else None,
        end_date=end_date.isoformat() if end_date else None,
        source=source_filter if source_filter != "All" else None,
        search_payee=search_payee or None,
        search=search_all or None,
        status=status_filter if status_filter != "All" else None,
    )

    if not rows:
        st.info("No transactions match the filters.")
        return

    # --- Build dataframe ---
    row_ids = []
    data = []
    for r in rows:
        row_ids.append(r["id"])
        data.append({
            "Date": r["date"],
            "Source": r["source"],
            "Payee": r["payee"] or "",
            "Category": r["category"] or "",
            "Subcategory": r["subcategory"] or "",
            "Amount": float(r["amount"]) if r["amount"] else 0.0,
            "Payor": r["payor"] or "",
            "Tax Flags": r["tax_flags"] or "",
            "Description (raw)": r["description_raw"],
            "Check #": r["check_number"] or "",
            "Note": r["note"] or "",
        })

    df = pd.DataFrame(data)

    # --- Summary ---
    total_out = df[df["Amount"] < 0]["Amount"].sum()
    total_in = df[df["Amount"] > 0]["Amount"].sum()
    col_a, col_b, col_c = st.columns(3)
    col_a.metric("Transactions", len(df))
    col_b.metric("Money Out", f"-${abs(total_out):,.2f}")
    col_c.metric("Money In", f"${total_in:,.2f}")

    # --- Export buttons ---
    col_x, col_y, _ = st.columns([1, 1, 4])
    with col_x:
        excel_buf = _df_to_excel(df)
        st.download_button(
            "Download Excel",
            data=excel_buf,
            file_name="abacus_browse.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col_y:
        csv_data = df.to_csv(index=False).encode("utf-8")
        st.download_button(
            "Download CSV (Print)",
            data=csv_data,
            file_name="abacus_browse.csv",
            mime="text/csv",
        )

    # --- Sortable, scrollable table with row selection ---
    # Apply consistent UI convention: credits in green, signed currency format.
    from ui._amount_style import styler_for_amount_column
    selection = st.dataframe(
        styler_for_amount_column(df, "Amount"),
        use_container_width=True,
        hide_index=True,
        height=500,
        on_select="rerun",
        selection_mode="single-row",
        key="browse_table",
    )

    # --- Edit selected transaction ---
    selected_rows = selection.get("selection", {}).get("rows", [])
    if selected_rows:
        idx = selected_rows[0]
        if idx < len(row_ids):
            txn_id = row_ids[idx]
            sel = data[idx]
            st.divider()
            st.markdown(f"**{sel['Date']}** | {sel['Payee']} | {sel['Source']} | {_fmt_browse_amt(sel['Amount'])}")

            col1, col2 = st.columns([3, 1])
            new_note = col1.text_input("Note", value=sel["Note"], key=f"browse_note_{txn_id}")
            if col2.button("Save", key=f"browse_save_{txn_id}"):
                queries.update_transaction(conn, txn_id, note=new_note if new_note else None)
                st.success("Note saved.")
                st.rerun()


def _fmt_browse_amt(val):
    if val > 0:
        # Credit — render in green via Streamlit markdown color span
        return f":green[\\${val:,.2f}]"
    if val < 0:
        return f"-${abs(val):,.2f}"
    return f"${val:,.2f}"


def _df_to_excel(df: pd.DataFrame) -> bytes:
    """Convert a DataFrame to an Excel file in memory."""
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Browse Results"

    # Header
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = openpyxl.styles.Font(bold=True)

    # Data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-filter and column widths
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col_name)) + 4, 12)

    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
